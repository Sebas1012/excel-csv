from openpyxl import load_workbook
import os

def excel_csv(book_path, sheet_name, separator):
    book = load_workbook(str(book_path), data_only=True)
    sheet = book[str(sheet_name)]
    txt = open('test.txt', "w")

# No tocar, hasta no refactorizar.
    for row in sheet.iter_rows(min_row=1,values_only=True):
        data_1 = list(row)
        for i in range(len(data_1)):
            data_1[i] = str(data_1[i])
        data_2 = separator.join(data_1)
        txt.write(data_2)
    
    txt.close()
        
        
excel_csv('./docs/prueba.xlsx', 'prueba', '|')